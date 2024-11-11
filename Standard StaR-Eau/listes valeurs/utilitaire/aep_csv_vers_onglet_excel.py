import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import os

# Dossier contenant les fichiers CSV - chemin à changer
dossier_csv = '.../Reseaux-eaux/Standard StaR-Eau/listes valeurs/brute avant import/'

# Nom du fichier Excel de sortie
fichier_excel = 'liste_valeurs_AEP.xlsx'

# Créer un nouveau classeur Excel
workbook = openpyxl.Workbook()
workbook.remove(workbook.active)  # Supprimer la feuille par défaut

# En-tête fixe
en_tete = ["code", "valeur", "description"]

# Parcourir tous les fichiers CSV dans le dossier
for fichier in os.listdir(dossier_csv):
    if fichier.startswith('aep_') and fichier.endswith('.csv'):
        # Lire le fichier CSV
        chemin_fichier = os.path.join(dossier_csv, fichier)
        df = pd.read_csv(chemin_fichier, header=None)
        
        # Nom de l'onglet (sans l'extension .csv)
        nom_onglet = os.path.splitext(fichier)[0]
        
        # Créer un nouvel onglet
        worksheet = workbook.create_sheet(title=nom_onglet)

        # Ajouter l'en-tête fixe
        worksheet.append(en_tete)

        # Ajouter les données du CSV
        for r in dataframe_to_rows(df, index=False, header=False):
            worksheet.append(r)

        # Formater l'en-tête
        for cell in worksheet[1]:
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
# Sauvegarder le fichier Excel
workbook.save(fichier_excel)
